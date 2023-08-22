import os
import re
import argparse
import random
import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt


def learn_openpyxl_deal_excel(fileName):
    # https://openpyxl.readthedocs.io/en/stable/index.html
    # 1 读取文件
    wb = openpyxl.load_workbook(fileName)
    sheet = wb['Sheet1']
    for sheet in wb:  # 遍历所有 sheet
        print(sheet.title)
    print(wb.sheetnames)

    # 2 获取单元格值
    # 1) 指定坐标范围的值
    cellss = sheet['A1:B5']
    # 2) 指定列的值
    cells = sheet['A']
    cellss = sheet['A:C']
    # 3) 指定行的值
    cells = sheet[5]
    cellss = sheet[5:7]
    # 4) 获取单元格的值 # 行下标从 1 开始 列下标从 0 开始
    print(sheet[1][0].value)
    # for cells in cellss:
        # for cell in cells:
            # print(cell.value)

    # 3 写入数据
    cell = sheet['D4']
    cell.value = '521'
    sheet.cell(1, 1).value = "write_Data"

    # 4 保存文件
    wb.save('data/new_data_openpyxl.xlsx')

    # 5 新建文件
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'newSheet'
    # 插入数据
    row = ["A", "B", "C"]
    worksheet.append(row)
    ws1 = workbook.create_sheet("Mysheet_End")  # insert at the end (default)
    ws2 = workbook.create_sheet("Mysheet_0", 0)  # insert at first position
    ws3 = workbook.create_sheet("Mysheet_pen", -1)  # insert at the penultimate position
    workbook.save('data/new_data_openpyxl_2.xlsx')
    workbook.close()


def learn_pandas_deal_excel(fileName):
    # 1 读取文件的同时必须指定工作表：
    sheet = pd.read_excel(fileName, sheet_name='Sheet1', index_col=False)

    # 2 获取单元格值
    # 第一行为标题行，所以从第二行才开始是其数据的第一行(idex=0)
    # print(sheet.head(2))
    # 1) 指定行的值 loc 根据所定义的index来获取行
    # print(sheet.loc[1])
    # print(sheet.iloc[1])
    # 2) 指定列的值
    print(sheet.iloc[:, 0]) # 列下标从 0 开始
    # 3) 获取单元格的值
    # print(sheet.loc[0][2])

    # 3 保存文件
    df = pd.DataFrame([1, 2, 3])
    df.to_excel("data/new_data_pandas.xlsx")


if __name__ == "__main__":
    xls_path = 'data/data.xls'
    xlsx_path = 'data/data.xlsx'

    learn_openpyxl_deal_excel(xlsx_path)
    learn_pandas_deal_excel(xls_path)
    learn_pandas_deal_excel(xlsx_path)
